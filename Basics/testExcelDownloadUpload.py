import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
filePath="C://Users//smitha.a.rao//Downloads//download.xlsx"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
#download excel
driver.find_element(By.CSS_SELECTOR,".button").click()
Dict ={}

def excel_update(filePath,fruitName,columnName,newValue):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    for i in range (1, sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == fruitName:
                Dict['rowPosition']=i
            if sheet.cell(row=i,column=j).value == columnName:
                Dict['columnPosition']=j
    print(Dict)
    sheet.cell(row=Dict['rowPosition'],column=Dict['columnPosition']).value = newValue
    print(sheet.cell(Dict['rowPosition'],column=Dict['columnPosition']).value)
    book.save(filePath)


fruitName="Apple"
columnName="price"
newValue =1000
# Call function to update excel data
excel_update(filePath,fruitName,columnName,newValue)


# Upload excel
uploadButton= driver.find_element(By.XPATH,"//input[@type='file']")
uploadButton.send_keys(filePath)

#Validate Toast
toast = (By.XPATH,"//div[@role='alert']/div[2]")
WebDriverWait(driver,10).until(EC.visibility_of_element_located(toast))
successMessage = driver.find_element(*toast).text
print(successMessage)
assert successMessage == "Updated Excel Data Successfully."

priceColumnID=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
fruitPrice = driver.find_element(By.XPATH,"//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+priceColumnID+"-undefined']").text
print(fruitName, fruitPrice)
assert int(fruitPrice) == newValue