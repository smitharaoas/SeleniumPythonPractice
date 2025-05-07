import openpyxl

workbook = openpyxl.load_workbook('../../../../Downloads/Pythonexcelpractice.xlsx')
sheet = workbook.active
# cell = sheet.cell(row=1,column=2)
# print(cell.value)
personalDetails =["Smitha", "Rao","F"]
Dict = {}
#To print table
for i in range(1 , sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Test4":
        for j in range(2,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
    k = 0
    if sheet.cell(row=i,column=1).value == "Test6":
        for j in range(2,sheet.max_column+1):
            sheet.cell(row=i,column=j).value = personalDetails[k]
            print(sheet.cell(row=i,column=j).value)
            k=k+1
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)